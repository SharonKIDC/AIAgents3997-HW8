"""Excel database manager for tenant management system.

Provides read/write operations for the Excel-based data store.
Uses openpyxl for Excel file manipulation with proper locking.
"""

import os
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional, Dict, Any
import json

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.config import get_config
from src.exceptions import DatabaseError, NotFoundError
from src.database.models import Tenant, TenantHistory, Building
from src.database.validators import DataValidator


class ExcelManager:
    """Manager for Excel-based tenant database."""

    TENANTS_SHEET = "Tenants"
    HISTORY_SHEET = "History"
    BUILDINGS_SHEET = "Buildings"

    TENANT_HEADERS = [
        "building_number", "apartment_number", "first_name", "last_name",
        "phone", "storage_number", "parking_slot_1", "parking_slot_2",
        "is_owner", "owner_first_name", "owner_last_name", "owner_phone",
        "whatsapp_members", "parking_authorizations", "move_in_date",
        "move_out_date", "palgate_access", "whatsapp_group"
    ]

    HISTORY_HEADERS = [
        "building_number", "apartment_number", "first_name", "last_name",
        "phone", "move_in_date", "move_out_date", "was_owner",
        "owner_first_name", "owner_last_name", "owner_phone"
    ]

    def __init__(self, db_path: str = None):
        """Initialize Excel manager with database path."""
        config = get_config()
        self._db_path = db_path or config.get("database.path", "data/tenants.xlsx")
        self._validator = DataValidator()
        self._ensure_database_exists()

    def _ensure_database_exists(self) -> None:
        """Create database file if it doesn't exist."""
        path = Path(self._db_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        if not path.exists():
            self._create_empty_database()

    def _create_empty_database(self) -> None:
        """Create empty database with proper structure."""
        wb = Workbook()
        ws_tenants = wb.active
        ws_tenants.title = self.TENANTS_SHEET
        for col, header in enumerate(self.TENANT_HEADERS, 1):
            ws_tenants.cell(row=1, column=col, value=header)
        ws_history = wb.create_sheet(self.HISTORY_SHEET)
        for col, header in enumerate(self.HISTORY_HEADERS, 1):
            ws_history.cell(row=1, column=col, value=header)
        ws_buildings = wb.create_sheet(self.BUILDINGS_SHEET)
        ws_buildings.cell(row=1, column=1, value="number")
        ws_buildings.cell(row=1, column=2, value="total_apartments")
        for building in Building.get_all_buildings():
            row = ws_buildings.max_row + 1
            ws_buildings.cell(row=row, column=1, value=building.number)
            ws_buildings.cell(row=row, column=2, value=building.total_apartments)
        wb.save(self._db_path)

    def _load_workbook(self) -> Workbook:
        """Load workbook with error handling."""
        try:
            return load_workbook(self._db_path)
        except Exception as e:
            raise DatabaseError(f"Failed to load database: {e}")

    def _save_workbook(self, wb: Workbook) -> None:
        """Save workbook with error handling."""
        try:
            wb.save(self._db_path)
        except Exception as e:
            raise DatabaseError(f"Failed to save database: {e}")

    def _date_to_str(self, d: date) -> str:
        """Convert date to string for storage."""
        return d.isoformat() if d else ""

    def _str_to_date(self, s: str) -> Optional[date]:
        """Convert string to date from storage."""
        if not s:
            return None
        if isinstance(s, datetime):
            return s.date()
        if isinstance(s, date):
            return s
        return date.fromisoformat(str(s))

    def get_tenant(self, building: int, apartment: int) -> Optional[Tenant]:
        """Get active tenant for an apartment."""
        wb = self._load_workbook()
        ws = wb[self.TENANTS_SHEET]
        for row in range(2, ws.max_row + 1):
            if (ws.cell(row, 1).value == building and
                ws.cell(row, 2).value == apartment and
                not ws.cell(row, 16).value):
                return self._row_to_tenant(ws, row)
        return None

    def _row_to_tenant(self, ws: Worksheet, row: int) -> Tenant:
        """Convert worksheet row to Tenant model."""
        from src.database.models import OwnerInfo
        owner_info = None
        if not ws.cell(row, 9).value:
            owner_info = OwnerInfo(
                first_name=ws.cell(row, 10).value or "",
                last_name=ws.cell(row, 11).value or "",
                phone=ws.cell(row, 12).value or ""
            )
        return Tenant(
            building_number=ws.cell(row, 1).value,
            apartment_number=ws.cell(row, 2).value,
            first_name=ws.cell(row, 3).value,
            last_name=ws.cell(row, 4).value,
            phone=ws.cell(row, 5).value,
            storage_number=ws.cell(row, 6).value,
            parking_slot_1=ws.cell(row, 7).value,
            parking_slot_2=ws.cell(row, 8).value,
            is_owner=bool(ws.cell(row, 9).value),
            owner_info=owner_info,
            whatsapp_members=json.loads(ws.cell(row, 13).value or "[]"),
            parking_authorizations=json.loads(ws.cell(row, 14).value or "[]"),
            move_in_date=self._str_to_date(ws.cell(row, 15).value),
            move_out_date=self._str_to_date(ws.cell(row, 16).value),
            palgate_access_enabled=bool(ws.cell(row, 17).value),
            whatsapp_group_enabled=bool(ws.cell(row, 18).value)
        )
