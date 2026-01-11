"""Excel database operations for tenant management.

Extended operations for the Excel-based data store including
create, update, delete, and history management.
"""

from datetime import date
from typing import List
import json

from openpyxl import load_workbook

from src.config import get_config
from src.exceptions import DatabaseError, NotFoundError, ValidationError
from src.database.models import Tenant, TenantHistory
from src.database.validators import DataValidator


class ExcelOperations:
    """Extended operations for Excel tenant database."""

    TENANTS_SHEET = "Tenants"
    HISTORY_SHEET = "History"

    def __init__(self, db_path: str = None):
        """Initialize with database path."""
        config = get_config()
        self._db_path = db_path or config.get("database.path", "data/tenants.xlsx")
        self._validator = DataValidator()

    def _load_workbook(self):
        """Load workbook with error handling."""
        try:
            return load_workbook(self._db_path)
        except Exception as e:
            raise DatabaseError(f"Failed to load database: {e}") from e

    def _save_workbook(self, wb) -> None:
        """Save workbook with error handling."""
        try:
            wb.save(self._db_path)
        except Exception as e:
            raise DatabaseError(f"Failed to save database: {e}") from e

    def create_tenant(self, tenant: Tenant) -> Tenant:
        """Create a new tenant record."""
        self._validator.validate_building_number(tenant.building_number)
        self._validator.validate_apartment_number(tenant.building_number, tenant.apartment_number)
        wb = self._load_workbook()
        ws = wb[self.TENANTS_SHEET]
        for row in range(2, ws.max_row + 1):
            if (
                ws.cell(row, 1).value == tenant.building_number
                and ws.cell(row, 2).value == tenant.apartment_number
                and not ws.cell(row, 16).value
            ):
                raise ValidationError(
                    "Active tenant already exists for this apartment",
                    {"building": tenant.building_number, "apartment": tenant.apartment_number},
                )
        row = ws.max_row + 1
        self._write_tenant_row(ws, row, tenant)
        self._save_workbook(wb)
        return tenant

    def update_tenant(self, tenant: Tenant) -> Tenant:
        """Update existing tenant record."""
        wb = self._load_workbook()
        ws = wb[self.TENANTS_SHEET]
        for row in range(2, ws.max_row + 1):
            if (
                ws.cell(row, 1).value == tenant.building_number
                and ws.cell(row, 2).value == tenant.apartment_number
                and not ws.cell(row, 16).value
            ):
                self._write_tenant_row(ws, row, tenant)
                self._save_workbook(wb)
                return tenant
        raise NotFoundError(
            "Tenant not found",
            {"building": tenant.building_number, "apartment": tenant.apartment_number},
        )

    def end_tenancy(self, building: int, apartment: int, move_out: date) -> TenantHistory:
        """End tenancy and move record to history."""
        wb = self._load_workbook()
        ws = wb[self.TENANTS_SHEET]
        for row in range(2, ws.max_row + 1):
            if (
                ws.cell(row, 1).value == building
                and ws.cell(row, 2).value == apartment
                and not ws.cell(row, 16).value
            ):
                ws.cell(row, 16).value = move_out.isoformat()
                history = self._create_history_record(ws, row, move_out)
                ws_history = wb[self.HISTORY_SHEET]
                self._write_history_row(ws_history, history)
                self._save_workbook(wb)
                return history
        raise NotFoundError("Active tenant not found")

    def _write_tenant_row(self, ws, row: int, tenant: Tenant) -> None:
        """Write tenant data to worksheet row."""
        ws.cell(row, 1).value = tenant.building_number
        ws.cell(row, 2).value = tenant.apartment_number
        ws.cell(row, 3).value = tenant.first_name
        ws.cell(row, 4).value = tenant.last_name
        ws.cell(row, 5).value = tenant.phone
        ws.cell(row, 6).value = tenant.storage_number
        ws.cell(row, 7).value = tenant.parking_slot_1
        ws.cell(row, 8).value = tenant.parking_slot_2
        ws.cell(row, 9).value = tenant.is_owner
        if tenant.owner_info:
            ws.cell(row, 10).value = tenant.owner_info.first_name
            ws.cell(row, 11).value = tenant.owner_info.last_name
            ws.cell(row, 12).value = tenant.owner_info.phone
        ws.cell(row, 13).value = json.dumps([m.model_dump() for m in tenant.whatsapp_members])
        ws.cell(row, 14).value = json.dumps([p.model_dump() for p in tenant.parking_authorizations])
        ws.cell(row, 15).value = tenant.move_in_date.isoformat()
        ws.cell(row, 16).value = tenant.move_out_date.isoformat() if tenant.move_out_date else None
        ws.cell(row, 17).value = tenant.palgate_access_enabled
        ws.cell(row, 18).value = tenant.whatsapp_group_enabled

    def _create_history_record(self, ws, row: int, move_out: date) -> TenantHistory:
        """Create history record from tenant row."""
        return TenantHistory(
            building_number=ws.cell(row, 1).value,
            apartment_number=ws.cell(row, 2).value,
            first_name=ws.cell(row, 3).value,
            last_name=ws.cell(row, 4).value,
            phone=ws.cell(row, 5).value,
            move_in_date=date.fromisoformat(str(ws.cell(row, 15).value)),
            move_out_date=move_out,
            was_owner=bool(ws.cell(row, 9).value),
            owner_first_name=ws.cell(row, 10).value,
            owner_last_name=ws.cell(row, 11).value,
            owner_phone=ws.cell(row, 12).value,
        )

    def _write_history_row(self, ws, history: TenantHistory) -> None:
        """Write history record to worksheet."""
        row = ws.max_row + 1
        ws.cell(row, 1).value = history.building_number
        ws.cell(row, 2).value = history.apartment_number
        ws.cell(row, 3).value = history.first_name
        ws.cell(row, 4).value = history.last_name
        ws.cell(row, 5).value = history.phone
        ws.cell(row, 6).value = history.move_in_date.isoformat()
        ws.cell(row, 7).value = history.move_out_date.isoformat()
        ws.cell(row, 8).value = history.was_owner
        ws.cell(row, 9).value = history.owner_first_name
        ws.cell(row, 10).value = history.owner_last_name
        ws.cell(row, 11).value = history.owner_phone

    def get_all_tenants(self, building: int = None) -> List[Tenant]:
        """Get all active tenants, optionally filtered by building."""
        from src.database.queries import TenantQueries

        queries = TenantQueries(self._db_path)
        return queries.get_all_tenants(building)

    def get_tenant_history(self, building: int, apartment: int) -> List[TenantHistory]:
        """Get history of all tenants for an apartment."""
        from src.database.history_manager import HistoryManager

        history_mgr = HistoryManager(self._db_path)
        return history_mgr.get_apartment_history(building, apartment)
