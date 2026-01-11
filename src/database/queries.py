"""Query operations for tenant database.

Provides read-only query operations for retrieving tenant data
from the Excel database without modifications.
"""

from typing import List, Optional

from openpyxl import load_workbook

from src.config import get_config
from src.exceptions import DatabaseError
from src.database.models import Tenant, Building


class TenantQueries:
    """Query operations for tenant data."""

    TENANTS_SHEET = "Tenants"

    def __init__(self, db_path: str = None):
        """Initialize with database path."""
        config = get_config()
        self._db_path = db_path or config.get("database.path", "data/tenants.xlsx")

    def _load_workbook(self):
        """Load workbook with error handling."""
        try:
            return load_workbook(self._db_path)
        except Exception as e:
            raise DatabaseError(f"Failed to load database: {e}") from e

    def get_all_tenants(self, building: int = None) -> List[Tenant]:
        """Get all active tenants, optionally filtered by building."""
        from src.database.excel_manager import ExcelManager
        manager = ExcelManager(self._db_path)
        workbook = self._load_workbook()
        worksheet = workbook[self.TENANTS_SHEET]
        tenants = []
        for row in range(2, worksheet.max_row + 1):
            if not worksheet.cell(row, 16).value:
                if building is None or worksheet.cell(row, 1).value == building:
                    tenants.append(manager._row_to_tenant(worksheet, row))
        return tenants

    def get_tenant(self, building: int, apartment: int) -> Optional[Tenant]:
        """Get active tenant for a specific apartment."""
        from src.database.excel_manager import ExcelManager
        manager = ExcelManager(self._db_path)
        return manager.get_tenant(building, apartment)

    def get_building_occupancy(self, building: int) -> dict:
        """Get occupancy statistics for a building."""
        building_info = Building.get_building(building)
        if not building_info:
            return {"error": f"Building {building} not found"}
        tenants = self.get_all_tenants(building)
        occupied = len(tenants)
        total = building_info.total_apartments
        return {
            "building": building,
            "total_apartments": total,
            "occupied": occupied,
            "vacant": total - occupied,
            "occupancy_rate": round(occupied / total * 100, 1) if total > 0 else 0
        }

    def get_all_buildings_occupancy(self) -> List[dict]:
        """Get occupancy statistics for all buildings."""
        buildings = Building.get_all_buildings()
        return [self.get_building_occupancy(b.number) for b in buildings]

    def search_tenants(
        self, name: str = None, phone: str = None, building: int = None
    ) -> List[Tenant]:
        """Search tenants by name, phone, or building."""
        tenants = self.get_all_tenants(building)
        results = []
        for tenant in tenants:
            if name:
                full_name = tenant.full_name.lower()
                if name.lower() not in full_name:
                    continue
            if phone:
                if phone not in tenant.phone:
                    continue
            results.append(tenant)
        return results

    def get_whatsapp_contacts(self, building: int = None) -> List[dict]:
        """Get phone numbers for WhatsApp group management."""
        tenants = self.get_all_tenants(building)
        contacts = []
        for tenant in tenants:
            if tenant.whatsapp_group_enabled:
                contacts.append({
                    "name": tenant.full_name,
                    "phone": tenant.phone,
                    "building": tenant.building_number,
                    "apartment": tenant.apartment_number
                })
                for member in tenant.whatsapp_members:
                    contacts.append({
                        "name": f"{member.first_name} {member.last_name}",
                        "phone": member.phone,
                        "building": tenant.building_number,
                        "apartment": tenant.apartment_number
                    })
        return contacts

    def get_parking_authorizations(self, building: int = None) -> List[dict]:
        """Get parking authorization list."""
        tenants = self.get_all_tenants(building)
        authorizations = []
        for tenant in tenants:
            if tenant.palgate_access_enabled:
                authorizations.append({
                    "name": tenant.full_name,
                    "phone": tenant.phone,
                    "building": tenant.building_number,
                    "apartment": tenant.apartment_number,
                    "slots": [tenant.parking_slot_1, tenant.parking_slot_2]
                })
                for auth in tenant.parking_authorizations:
                    authorizations.append({
                        "name": f"{auth.first_name} {auth.last_name}",
                        "phone": auth.phone,
                        "building": tenant.building_number,
                        "apartment": tenant.apartment_number,
                        "slots": []
                    })
        return authorizations
