"""History management for tenant records.

Provides operations for managing tenant history records,
including creating history entries and querying past tenants.
"""

from datetime import date
from typing import List

from openpyxl import load_workbook

from src.config import get_config
from src.exceptions import DatabaseError
from src.database.models import TenantHistory


class HistoryManager:
    """Manager for tenant history records."""

    HISTORY_SHEET = "History"

    def __init__(self, db_path: str = None):
        """Initialize with database path."""
        config = get_config()
        self._db_path = db_path or config.get("database.path", "data/tenants.xlsx")

    def _load_workbook(self):
        """Load workbook with error handling."""
        try:
            return load_workbook(self._db_path)
        except Exception as e:
            raise DatabaseError(f"Failed to load database: {e}") from e

    def _save_workbook(self, workbook) -> None:
        """Save workbook with error handling."""
        try:
            workbook.save(self._db_path)
        except Exception as e:
            raise DatabaseError(f"Failed to save database: {e}") from e

    def add_history_record(self, history: TenantHistory) -> TenantHistory:
        """Add a history record to the database."""
        workbook = self._load_workbook()
        worksheet = workbook[self.HISTORY_SHEET]
        self._write_history_row(worksheet, history)
        self._save_workbook(workbook)
        return history

    def _write_history_row(self, worksheet, history: TenantHistory) -> None:
        """Write history record to worksheet."""
        row = worksheet.max_row + 1
        worksheet.cell(row, 1).value = history.building_number
        worksheet.cell(row, 2).value = history.apartment_number
        worksheet.cell(row, 3).value = history.first_name
        worksheet.cell(row, 4).value = history.last_name
        worksheet.cell(row, 5).value = history.phone
        worksheet.cell(row, 6).value = history.move_in_date.isoformat()
        worksheet.cell(row, 7).value = history.move_out_date.isoformat()
        worksheet.cell(row, 8).value = history.was_owner
        worksheet.cell(row, 9).value = history.owner_first_name
        worksheet.cell(row, 10).value = history.owner_last_name
        worksheet.cell(row, 11).value = history.owner_phone

    def get_apartment_history(self, building: int, apartment: int) -> List[TenantHistory]:
        """Get history of all tenants for an apartment."""
        workbook = self._load_workbook()
        worksheet = workbook[self.HISTORY_SHEET]
        history = []
        for row in range(2, worksheet.max_row + 1):
            if (
                worksheet.cell(row, 1).value == building
                and worksheet.cell(row, 2).value == apartment
            ):
                history.append(self._row_to_history(worksheet, row))
        return sorted(history, key=lambda h: h.move_in_date, reverse=True)

    def get_building_history(self, building: int) -> List[TenantHistory]:
        """Get history of all tenants for a building."""
        workbook = self._load_workbook()
        worksheet = workbook[self.HISTORY_SHEET]
        history = []
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row, 1).value == building:
                history.append(self._row_to_history(worksheet, row))
        return sorted(history, key=lambda h: h.move_in_date, reverse=True)

    def _row_to_history(self, worksheet, row: int) -> TenantHistory:
        """Convert worksheet row to TenantHistory model."""
        return TenantHistory(
            building_number=worksheet.cell(row, 1).value,
            apartment_number=worksheet.cell(row, 2).value,
            first_name=worksheet.cell(row, 3).value,
            last_name=worksheet.cell(row, 4).value,
            phone=worksheet.cell(row, 5).value,
            move_in_date=date.fromisoformat(str(worksheet.cell(row, 6).value)),
            move_out_date=date.fromisoformat(str(worksheet.cell(row, 7).value)),
            was_owner=bool(worksheet.cell(row, 8).value),
            owner_first_name=worksheet.cell(row, 9).value,
            owner_last_name=worksheet.cell(row, 10).value,
            owner_phone=worksheet.cell(row, 11).value,
        )
